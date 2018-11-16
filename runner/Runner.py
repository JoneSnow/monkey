# -*- coding:utf-8 -*-
# Auther: guofengyang
# Date: 2018/8/24 11:46
import copy
import datetime
import gzip
import logging
import os
import shutil
import subprocess
import sys
import time
from itertools import islice
from multiprocessing.pool import ThreadPool as Pool

from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles import colors, PatternFill, Border, Side, Alignment
from openpyxl.styles import numbers

# 添加包路径
path = os.path.dirname(os.path.dirname(os.path.abspath(os.path.realpath(__file__))))
sys.path.append(path)
from runner import RUNNER, ROOT
from runner.log import init_log
from runner.tools import Tools

logger = logging.getLogger(__file__)
import os.path

class Runner(object):
    def __init__(self):
        self.config = Tools.parseConfigJson()
        self.result = os.path.join(ROOT, "result")

    def init(self):
        result = self.result
        # if os.path.exists(result):
        #     shutil.rmtree(result)
        #     time.sleep(1)
        # create log path
        if not os.path.exists(result):
            os.mkdir(result)
            time.sleep(1)
        # logger初始化
        init_log()

    def run(self):
        """
        入口函数
        :return:{u'0123456789ABCDEF': {'duration': '0:00:04.237000', 'crash': 2, 'detail': {'01-01 00:12:12.080': ' *** FATAL EXCEPTION IN SYSTEM PROCESS: main\r\n java.lang.NullPointerException\r\n \tat com.android.commands.monkey.MonkeySourceRandom.randomPoint(MonkeySourceRandom.java:324)\r\n', '01-01 00:12:11.070': ' *** FATAL EXCEPTION IN SYSTEM PROCESS: UI\r\n java.lang.NullPointerException\r\n \tat com.android.internal.widget.PointerLocationView.addPointerEvent(PointerLocationView.java:552)\r\n'}, 'anr': 0, 'rom': '03.02.08950.H30.00009'}}
        :rtype:dict
        """
        logger.debug(u"start")
        info = []
        result = {}
        Tools.execute("adb kill-server")
        Tools.execute("adb start-server")
        self.checkDevices()
        self.init()
        pool = Pool(processes=len(self.config))
        for key, value in self.config.items():
            sn = value["sn"]
            packages = value["packages"]
            throttle = value["throttle"]
            info.append(pool.apply_async(self.monkey, (sn, packages, throttle,)))
        pool.close()
        pool.join()
        for i in info:
            result.update(i.get())
        self.to_html(result)
        return result

    def checkDevices(self):
        """
        检查配置文件中的sn号是否存在于已连接的设备中
        :raise: ValueError
        """
        # 连接网络设备
        logger.debug(u"链接网络设备")
        config_devices = self.getConfigDevices()
        for device in config_devices:
            if ":" in device:
                Tools.execute("adb connect {ip}".format(ip=device))
        logger.debug(u"检查配置文件中的设备是否在链接状态")
        results = []
        connect_devices = Tools.devices()
        config_devices = self.getConfigDevices()
        for item in config_devices:
            if item not in connect_devices:
                results.append(item)
        if results:
            logger.error(u"检查设备链接状态失败", exc_info=True)
            raise ValueError(
                u"can't find config devices {} in connect devices, please check config.json".format(results))

    def getConfigDevices(self):
        '''
        获取config配置文件的所有sn号
        :return: 返回含有sn号的生成器
        '''
        sns = [value["sn"] for key, value in self.config.items()]
        logger.debug(u"获取配置文件的设备号{}".format(sns))
        return sns

    def monkey(self, sn, packages, throttle):
        logger.debug(u"{}, monkey测试开始".format(sn))
        sn_string = sn.replace(":", "_") if ":" in sn else sn
        info = {}
        package_string = ""
        starttime = datetime.datetime.now()
        t = starttime.strftime("%Y%m%d%H%M%S")
        logcat_log = "logcat_{}_{}".format(sn_string, t)
        monkey_log = "monkey_{}_{}".format(sn_string, t)
        result_path = os.path.join(ROOT, "result")
        log_path = os.path.join(result_path, sn_string)
        if not os.path.exists(log_path):
            os.mkdir(log_path)
        logcat_log_path = os.path.join(log_path, logcat_log)
        monkey_log_path = os.path.join(log_path, monkey_log)
        # rom 版本号
        rom = Tools(sn).getSystemVersion()
        for package in packages[:]:
            package_string += "-p {} ".format(package)
        try:
            p1 = subprocess.Popen("adb -s {sn} logcat -c && adb -s {sn} logcat -v time > {logcat_log_path}".format(
                sn=sn,logcat_log_path=logcat_log_path),shell=True)
            Tools.execute(
                "adb -s {sn} shell monkey {package_string} --ignore-timeouts --ignore-crashes"
                " --ignore-security-exceptions --kill-process-after-error --monitor-native-crashes"
                " --pct-syskeys 0 -v -v -v -s 3343 --throttle {throttle} > {mpath}".format(sn=sn, package_string=package_string,
                                                                           throttle=throttle, mpath=monkey_log_path))
        finally:
            Tools.execute("taskkill /t /f /pid {}".format(p1.pid))
        #get log
        Tools.execute("adb -s {sn} pull /data/system/dropbox/ {log_path}".format(sn=sn, log_path=log_path))
        Tools.execute("adb -s {sn} pull /data/anr/ {log_path}".format(sn=sn, log_path=log_path))
        Tools.execute("adb -s {sn} pull /data/tombstone {log_path}".format(sn=sn, log_path=log_path))
        #prase log
        endtime = datetime.datetime.now()
        duration = endtime - starttime
        # parse log
        detail = []
        summary = { "anr":0, "crash":0, "strictmode":0, "other":0}
        detail_package_summary = {}
        for item in self.parse_log(log_path):
            detail.append(item)
            packagename = item.get("package", None)
            if packagename:
                # res = map(lambda x: x in detail_package_summary, packagename)
                # if not all(res):
                if packagename not in  detail_package_summary:
                    detail_package_summary.update({packagename:{"anr": 0, "crash":0, "strictmode":0, "other":0}})
                if item["anr"]:
                    detail_package_summary[packagename]["anr"] += 1
                elif item["crash"]:
                    detail_package_summary[packagename]["crash"] += 1
                elif item["strictmode"]:
                    detail_package_summary[packagename]["strictmode"] += 1
                elif item["other"]:
                    detail_package_summary[packagename]["other"] += 1
            if item["anr"]:
                summary["anr"] += 1
            elif item["crash"]:
                summary["crash"] += 1
            elif item["strictmode"]:
                summary["strictmode"] += 1
            elif item["other"]:
                summary["other"] += 1
        info[sn] = {"duration": duration, "rom": rom, "detail": detail, "summary": summary, "detail_package_summary": detail_package_summary }
        logger.debug(u"{}, monkey执行结果数据:{}".format(sn, info))
        return info

    def to_excel(self, dic):
        logger.debug(u"生成测试结果表格")
        filepath = os.path.join(self.result, "result_{}.xlsx".format(datetime.datetime.now().strftime("%Y%m%d%H%M%S")))
        template_path = os.path.join(RUNNER, "template.xlsx")
        dic = copy.deepcopy(dic)
        shutil.copy(template_path, filepath)
        wb = load_workbook(filepath)
        # 简介页面
        ws = wb["Summary"]
        start_row = 3
        font = Font(name=u"宋体", color=colors.BLACK, size=11)
        fill = PatternFill(fill_type="solid", bgColor="F1E6DC", fgColor="F1E6DC")
        side = Side(border_style="thin", color=colors.BLACK)
        border = Border(left=side, right=side, top=side, bottom=side, outline=side)
        for key, value in dic.items():
            # 行高
            ws.row_dimensions[start_row].height = 30
            for i in ["A", "B", "D", "C", "E", "F"]:
                ws["{}{}".format(i, start_row)].font = font
                ws["{}{}".format(i, start_row)].fill = fill
                ws["{}{}".format(i, start_row)].border = border
                # 设置时间格式为文本
                if i == "D":
                    ws["{}{}".format(i, start_row)].number_format = numbers.FORMAT_TEXT

            ws["A{}".format(start_row)] = value["rom"]
            ws["B{}".format(start_row)] = key
            ws["D{}".format(start_row)] = value["duration"][0:-3]
            ws["E{}".format(start_row)] = value["crash"]
            ws["F{}".format(start_row)] = value["anr"]
            if value["crash"] + value["anr"]:
                result = "FAIL"
                comFont = Font(size=11, bold=True, name=u"宋体", color=colors.RED)
                ws["C{}".format(start_row)].font = comFont
            else:
                result = "PASS"
            ws["C{}".format(start_row)] = result
            start_row += 1
        # 详细页面
        ws = wb["Detail"]
        start_row = 3
        fill2 = PatternFill(fill_type="solid", bgColor="DEF1EB", fgColor="DEF1EB")
        alignment = Alignment(wrap_text=True)
        for key, value in dic.items():
            detail = value["detail"]
            for detail_key, detail_value in detail.items():
                ws.row_dimensions[start_row].height = 50
                for i in ["A", "B", "C"]:

                    ws["{}{}".format(i, start_row)].fill = fill2
                    ws["{}{}".format(i, start_row)].border = border
                    if i == "C":
                        ws["{}{}".format(i, start_row)].alignment = alignment
                    if i == "B":
                        ws["{}{}".format(i, start_row)].number_format = numbers.FORMAT_TEXT
                ws["A{}".format(start_row)] = key
                ws["B{}".format(start_row)] = detail_key
                ws["C{}".format(start_row)] = detail_value.strip()
                start_row += 1
        wb.save(filepath)
        logger.debug(u"测试结束")

    def to_html(self, dic):
        from jinja2 import PackageLoader, Environment

        env = Environment(loader=PackageLoader("runner", 'templates'))  # 创建一个包加载器对象

        template = env.get_template('template.html')  # 获取一个模板文件
        s = template.render(dict = dic)
        with open("../result/report.html", "w") as f:
            f.write(s)

    def parse_log(self, path):
        dropbox = os.path.join(path, "dropbox")
        for filename in os.listdir(dropbox):
            filepath = os.path.join(dropbox, filename)
            yield self.get_info(filename, filepath)

    def get_info(self, filename, filepath):
        info = {
            "filename": filename,
            "anr": 0,
            "crash": 0,
            "strictmode": 0,
            "other": 0
        }
        if filename.startswith("system_app"):
            for item in [".txt", ".gz"]:
                detail = self.get_info_detail(filepath, item)
                if detail:
                    package, msg = detail
                    info["msg"] = msg
                    info["package"] = package
        if filename.startswith("system_app_anr"):
            info["anr"] += 1
        elif filename.startswith("system_app_crash"):
            info["crash"] += 1
        elif filename.startswith("system_app_strictmode"):
            info["strictmode"] += 1
        else:
            info["other"] += 1
        return info

    def get_info_detail(self, filepath, filetype=".txt"):
        if filepath.endswith(filetype):
            package = []
            if filetype == ".txt":
                with open(filepath, "r") as f:
                    msg = islice(f, 0, 10)
                    msg = list(msg)
                    for i in msg:
                        if "Package" in i:
                            package.append(i.split(":")[1].strip())
                    package = ",".join(package)
                    msg = "<br>".join(msg)
                    return (package, msg)
            elif filetype == ".gz":
                with gzip.open(filepath, "r") as f:
                    msg = islice(f, 0, 10)
                    msg = list(msg)
                    for i in msg:
                        if "Package" in i:
                            package.append(i.split(":")[1].strip())
                    package = ",".join(package)
                    msg = "<br>".join(msg)
                    return (package, msg)

if __name__ == "__main__":
    r = Runner()
    r.run()
